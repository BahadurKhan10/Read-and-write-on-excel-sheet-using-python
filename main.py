#CONNECTING AND READING EXCEL
from openpyxl import load_workbook
workbook = load_workbook (filename="excelwork1.xlsx") #file in same directory workbook.sheetnames
sheet workbook.active
sheet
#Sheet 1
sheet.title
#Output 'Sheet 1'



#cell info
print(sheet ["A1"])

#Printing values form sheet
print(sheet ["A1"].value)
print(sheet ["A2"].value)
print(sheet ["A3"].value)
print(sheet ["A4"].value)
print(sheet ["A5").value)
print(sheet ["A6"].value)

#Output
#ΝΑΜΕ
#ReXX
#Haris
#Bahadur
#Hamza
#Uzair



#INSERTING DATA OF SQL TO TABLE 
import sqlite3
con = sqlite3.connect('vpms.db') 
cursor = con.cursor() 
cur con.cursor()

# Create table
cur.execute('''CREATE TABLE EXCELWORKS
(NAME, ROLLNO)''')

#Insert
cur.execute("INSERT INTO EXCELWORKS VALUES ('MESSI',10)")
con.commit()
con.close()